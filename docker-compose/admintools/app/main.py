import io
import os
import secrets
import string
import time

import openpyxl
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy import func

from . import ldap_client, sync
from .db import DEFAULT_DOMAIN, GroupMember, MailAccount, SessionLocal, init_db


def _gen_password() -> str:
    return "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(14)) + "Aa9!"

API_KEY = os.environ.get("ADMINTOOLS_API_KEY", "")


def require_api_key(x_api_key: str = Header(default="")):
    """Auth for the external read/write API (X-API-Key header)."""
    if not API_KEY:
        raise HTTPException(503, "API disabled: set ADMINTOOLS_API_KEY")
    if x_api_key != API_KEY:
        raise HTTPException(401, "invalid API key")


class AccountIn(BaseModel):
    username: str
    email: str = ""
    password: str = ""
    quota_gb: int = 5
    is_active: bool = True
    kind: str = "shared"
    notes: str = ""


class AccountPatch(BaseModel):
    email: str | None = None
    password: str | None = None
    quota_gb: int | None = None
    is_active: bool | None = None
    kind: str | None = None
    notes: str | None = None


def _account_out(a: MailAccount) -> dict:
    """Public representation — never exposes the password."""
    return {"id": a.id, "username": a.username, "email": a.email, "kind": a.kind,
            "type": a.type, "quota_gb": a.quota_gb, "is_active": a.is_active, "notes": a.notes}


def _shared_type(username: str) -> str:
    if username.startswith("app-"):
        return "app"
    if username.startswith("grp-"):
        return "grp"
    return "other"


app = FastAPI(title="adminTools")
templates = Jinja2Templates(directory="app/templates")


def _parse_accounts_xlsx(data: bytes) -> list[dict]:
    """Parse mail accounts from an email-list xlsx across all sheets.
    Columns (by header): username, password, email (with domain), active, displayname."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    accounts: list[dict] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        def col(*names: str) -> int:
            for n in names:
                for i, h in enumerate(header):
                    if h == n:
                        return i
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return -1

        u_i, p_i, a_i, d_i = col("username"), col("password"), col("active"), col("displayname")
        e_i = col("email (with domain)")
        if e_i < 0:
            for i, h in enumerate(header):
                if "email" in h and "domain" in h:
                    e_i = i
                    break
        if e_i < 0:
            e_i = col("email")
        if u_i < 0 or e_i < 0:
            continue
        for row in rows[1:]:
            row = list(row) + [None] * (len(header) - len(row or []))
            username = str(row[u_i]).strip() if row[u_i] is not None else ""
            email = str(row[e_i]).strip() if row[e_i] is not None else ""
            if not username or "@" not in email:
                continue
            av = row[a_i] if a_i >= 0 else True
            is_active = av is True or (isinstance(av, str) and av.strip().lower() in
                                       ("true", "yes", "1", "on", "○", "〇", "x", "✓", "✔"))
            accounts.append({
                "username": username,
                "password": str(row[p_i]).strip() if p_i >= 0 and row[p_i] is not None else "",
                "email": email,
                "is_active": is_active,
                "notes": str(row[d_i]).strip() if d_i >= 0 and row[d_i] is not None else "",
            })
    return accounts


@app.post("/import")
def import_accounts(file: UploadFile = File(...), replace: bool = Form(False)):
    accounts = _parse_accounts_xlsx(file.file.read())
    # Categorize kind by AD membership (username in AD -> user, else shared).
    try:
        ad = {u["username"].lower() for u in ldap_client.list_users()}
    except Exception:  # noqa: BLE001
        ad = None
    usernames = {a["username"] for a in accounts}
    inserted = updated = removed = 0
    with SessionLocal() as s:
        for a in accounts:
            kind = "user" if (ad is not None and a["username"].lower() in ad) else "shared"
            atype = "user" if kind == "user" else _shared_type(a["username"])
            existing = s.query(MailAccount).filter_by(username=a["username"]).first()
            if existing:
                if a["password"]:
                    existing.password = a["password"]
                existing.email, existing.is_active, existing.notes, existing.kind, existing.type = (
                    a["email"], a["is_active"], a["notes"], kind, atype)
                updated += 1
            else:
                s.add(MailAccount(username=a["username"], password=a["password"] or _gen_password(),
                                  email=a["email"], quota_gb=5, is_active=a["is_active"],
                                  kind=kind, type=atype, notes=a["notes"]))
                inserted += 1
        if replace:
            removed = s.query(MailAccount).filter(MailAccount.username.notin_(usernames)).delete(
                synchronize_session=False)
        s.commit()
    return RedirectResponse(f"/email?imported={inserted}&updated={updated}&removed={removed}", status_code=303)


@app.on_event("startup")
def _startup() -> None:
    init_db()


@app.get("/healthz")
def healthz():
    return {"status": "ok"}


_ldap_cache: dict = {"users": None, "ts": 0.0, "groups": None}


def _ldap_users_cached() -> list[dict]:
    if _ldap_cache["users"] is None or time.time() - _ldap_cache["ts"] > 300:
        _ldap_cache["users"] = ldap_client.list_users()
        _ldap_cache["ts"] = time.time()
    return _ldap_cache["users"]


@app.get("/ldap/users")
def ldap_users_json():
    """AD users for the searchable username picker (cached 5 min)."""
    try:
        return {"users": _ldap_users_cached()}
    except Exception as e:  # noqa: BLE001
        return JSONResponse({"error": str(e)[:140], "users": []})


@app.get("/ldap/groups")
def ldap_groups_json():
    """AD groups for the bulk-import picker (cached 5 min)."""
    if _ldap_cache["groups"] is None or time.time() - _ldap_cache["ts"] > 300:
        try:
            _ldap_cache["groups"] = ldap_client.list_groups()
        except Exception as e:  # noqa: BLE001
            return JSONResponse({"error": str(e)[:140], "groups": []})
    return {"groups": _ldap_cache["groups"]}


@app.post("/ldap/import_group")
def ldap_import_group(group_dn: str = Form(...)):
    """Onboard a whole AD group: upsert each enabled member as a User mailbox,
    populating fields from AD (display name -> notes; AD mail -> alias if it
    differs from <username>@domain)."""
    try:
        members = ldap_client.group_members(group_dn)
    except Exception as e:  # noqa: BLE001
        return RedirectResponse(f"/email?err=LDAP エラー: {str(e)[:80]}#user", status_code=303)
    created = updated = 0
    with SessionLocal() as s:
        for u in members:
            if u["disabled"]:
                continue
            mail = (u["mail"] or "").strip()
            email = mail if (mail and mail.split("@")[0] != u["username"]) else f"{u['username']}@{DEFAULT_DOMAIN}"
            a = s.query(MailAccount).filter_by(username=u["username"]).first()
            if a:
                a.kind, a.type = "user", "user"
                if not a.notes and u["display_name"]:
                    a.notes = u["display_name"]
                updated += 1
            else:
                s.add(MailAccount(username=u["username"], password=_gen_password(), email=email,
                                  quota_gb=5, is_active=True, kind="user", type="user",
                                  notes=u["display_name"]))
                created += 1
        s.commit()
    return RedirectResponse(f"/email?imported={created}&updated={updated}&removed=0#user", status_code=303)


def _render_index(request: Request, active_app: str):
    with SessionLocal() as s:
        accounts = s.query(MailAccount).order_by(MailAccount.username).all()
        counts = dict(s.query(GroupMember.group_id, func.count()).group_by(GroupMember.group_id).all())
    return templates.TemplateResponse("index.html", {
        "request": request,
        "active_app": active_app,           # 'home' | 'email' | 'kot' — drives which view shows
        "shared": [a for a in accounts if a.kind != "user"],
        "user": [a for a in accounts if a.kind == "user"],
        "member_counts": counts,
        "sync": sync.get_state(),
        "domain": DEFAULT_DOMAIN,
    })


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return _render_index(request, "home")


@app.get("/email", response_class=HTMLResponse)
def app_email(request: Request):
    return _render_index(request, "email")


@app.get("/kot", response_class=HTMLResponse)
def app_kot(request: Request):
    return _render_index(request, "kot")


@app.get("/groups/{gid}/members")
def group_members_list(gid: int):
    with SessionLocal() as s:
        rows = s.query(GroupMember).filter_by(group_id=gid).order_by(GroupMember.username).all()
        return {"members": [m.username for m in rows]}


@app.post("/groups/{gid}/members")
def group_member_add(gid: int, username: str = Form(...)):
    username = username.strip()
    with SessionLocal() as s:
        g = s.get(MailAccount, gid)
        if not g or g.type != "grp":
            raise HTTPException(400, "not a group")
        if username and not s.query(GroupMember).filter_by(group_id=gid, username=username).first():
            s.add(GroupMember(group_id=gid, username=username))
            s.commit()
        rows = s.query(GroupMember).filter_by(group_id=gid).order_by(GroupMember.username).all()
        return {"members": [m.username for m in rows]}


@app.post("/groups/{gid}/members/remove")
def group_member_remove(gid: int, username: str = Form(...)):
    with SessionLocal() as s:
        s.query(GroupMember).filter_by(group_id=gid, username=username.strip()).delete()
        s.commit()
        rows = s.query(GroupMember).filter_by(group_id=gid).order_by(GroupMember.username).all()
        return {"members": [m.username for m in rows]}


def _alias_email(username: str, use_alias: bool, alias: str) -> str:
    """Email is the custom alias when enabled+provided, else <username>@domain."""
    return alias.strip() if (use_alias and alias.strip()) else f"{username}@{DEFAULT_DOMAIN}"


@app.post("/accounts")
def create_account(
    username: str = Form(""),
    name: str = Form(""),
    type: str = Form("other"),
    password: str = Form(""),
    quota_gb: int = Form(5),
    is_active: bool = Form(False),
    kind: str = Form("shared"),
    use_alias: bool = Form(False),
    alias: str = Form(""),
    notes: str = Form(""),
):
    kind = kind if kind in ("user", "shared") else "shared"
    if kind == "shared":
        # Shared ID is generated from type + name:  <type>-<name>.
        type = type if type in ("app", "grp", "other") else "other"
        name = name.strip()
        if not name:
            return RedirectResponse("/email?err=名前を入力してください", status_code=303)
        username = f"{type}-{name}"
    else:
        username = username.strip()
        type = "user"
        # Private (user) emails must map to a real AD account.
        try:
            if not ldap_client.find_user(username):
                return RedirectResponse(f"/email?err=AD に {username} が見つかりません#user", status_code=303)
        except Exception as e:  # noqa: BLE001
            return RedirectResponse(f"/email?err=LDAP エラー: {str(e)[:80]}#user", status_code=303)
    with SessionLocal() as s:
        if s.query(MailAccount).filter_by(username=username).first():
            return RedirectResponse(f"/email?err={username} は既に存在します", status_code=303)
        # Group/shared mailbox defaults to <groupname>@domain; alias addresses are
        # set separately (編集 → 追加エイリアス). Users keep the alias-on-create flow.
        email = f"{username}@{DEFAULT_DOMAIN}" if kind == "shared" else _alias_email(username, use_alias, alias)
        s.add(MailAccount(username=username, password=password or _gen_password(),
                          email=email, quota_gb=quota_gb,
                          is_active=is_active, kind=kind, type=type, notes=notes))
        s.commit()
    return RedirectResponse("/email#user" if kind == "user" else "/email", status_code=303)


@app.post("/accounts/{account_id}/update")
def update_account(
    account_id: int,
    password: str = Form(...),
    quota_gb: int = Form(5),
    is_active: bool = Form(False),
    use_alias: bool = Form(False),
    alias: str = Form(""),
    notes: str = Form(""),
    extra_aliases: str = Form(""),
):
    with SessionLocal() as s:
        a = s.get(MailAccount, account_id)
        if not a:
            raise HTTPException(404)
        if password and password != a.password:
            a.password_dirty = True  # changed in adminTools -> push to Sakura on next sync
        a.password, a.quota_gb, a.is_active, a.notes = password, quota_gb, is_active, notes
        # Group/shared mailbox is always <name>@domain; all its addresses are aliases.
        # Users keep the single-custom-alias (old address) behaviour.
        a.email = f"{a.username}@{DEFAULT_DOMAIN}" if a.kind == "shared" else _alias_email(a.username, use_alias, alias)
        # Alias addresses (one per line); normalize to newline-separated.
        a.extra_aliases = "\n".join(
            e.strip() for e in extra_aliases.replace(",", "\n").splitlines() if e.strip()
        )
        kind = a.kind
        s.commit()
    return RedirectResponse("/email#user" if kind == "user" else "/email", status_code=303)


@app.post("/accounts/{account_id}/delete")
def delete_account(account_id: int):
    with SessionLocal() as s:
        a = s.get(MailAccount, account_id)
        if a:
            s.delete(a)
            s.commit()
    return RedirectResponse("/email", status_code=303)


@app.get("/sync/preview")
def sync_preview():
    """The check state shown before syncing: planned user + alias creates."""
    return JSONResponse(sync.preview())


@app.post("/sync/check")
def sync_check_start():
    """Start a background live Sakura check (read-only). It logs in + reads the
    panel (~45s), so the UI starts it then polls /sync/check/status — avoiding a
    long-lived request that proxies/browsers would cut off."""
    started = sync.start_check(refresh=True)
    return JSONResponse({"started": started, **sync.get_check_state()})


@app.get("/sync/check/status")
def sync_check_status():
    """Poll the background Sakura check: {running, result, error}."""
    return JSONResponse(sync.get_check_state())


def _account_aliases(a: MailAccount) -> list[str]:
    """All alias addresses for an account: the primary email (if it differs from
    <username>@domain) plus extra_aliases."""
    out = []
    local = (a.email or "").split("@")[0]
    if local and local != a.username:
        out.append(a.email)
    out += [x.strip() for x in (a.extra_aliases or "").replace(",", "\n").splitlines() if x.strip()]
    # de-dup, preserve order
    seen, uniq = set(), []
    for e in out:
        if e not in seen:
            seen.add(e); uniq.append(e)
    return uniq


@app.get("/export/xlsx")
def export_xlsx(passwords: bool = False):
    """Download an Excel workbook: sheet 1 = user mailboxes (LDAP username, email,
    aliases); sheet 2 = group emails with assigned members (comma-joined).
    passwords=true adds a パスワード column (mailbox password; aliases have none)."""
    wb = openpyxl.Workbook()
    pw_hdr = ["パスワード"] if passwords else []
    with SessionLocal() as s:
        # Sheet 1 — user emails
        ws = wb.active
        ws.title = "ユーザーメール"
        ws.append(["LDAPユーザー名", "メールアドレス", "エイリアス"] + pw_hdr + ["表示名", "状態"])
        for a in s.query(MailAccount).filter(MailAccount.kind == "user").order_by(MailAccount.username).all():
            row = [a.username, f"{a.username}@{DEFAULT_DOMAIN}", ", ".join(_account_aliases(a))]
            if passwords:
                row.append(a.password or "")
            row += [a.notes or "", "有効" if a.is_active else "無効"]
            ws.append(row)

        # Sheet 2 — group emails + assigned members (comma-joined)
        wg = wb.create_sheet("グループメール")
        wg.append(["グループ", "受信箱", "エイリアス", "メンバー（カンマ区切り）"] + pw_hdr + ["種別", "状態"])
        for a in s.query(MailAccount).filter(MailAccount.kind != "user").order_by(MailAccount.username).all():
            members = [m.username for m in s.query(GroupMember).filter_by(group_id=a.id)
                       .order_by(GroupMember.username).all()]
            row = [a.username, f"{a.username}@{DEFAULT_DOMAIN}", ", ".join(_account_aliases(a)), ", ".join(members)]
            if passwords:
                row.append(a.password or "")
            row += [a.type, "有効" if a.is_active else "無効"]
            wg.append(row)

    for sheet in wb.worksheets:                       # widen columns a bit
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "email-list-with-passwords.xlsx" if passwords else "email-list.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.post("/sync")
def trigger_sync(remove: bool = Form(False), sync_all_passwords: bool = Form(False)):
    # remove deletes Sakura mailboxes not in the DB — OFF unless explicitly opted in.
    # sync_all_passwords pushes every password (slow); otherwise only changed ones.
    sync.start_sync(remove_not_on_list=remove, sync_all_passwords=sync_all_passwords)
    return RedirectResponse("/email", status_code=303)


@app.get("/sync/status")
def sync_status():
    return JSONResponse(sync.get_state())


@app.get("/imports")
def imports_log():
    """Recent King of Time import runs (read from the kot-import DB)."""
    url = os.environ.get("KOT_DB_URL", "")
    if not url:
        return {"enabled": False, "runs": []}
    try:
        import psycopg
        with psycopg.connect(url, connect_timeout=5) as c, c.cursor() as cur:
            cur.execute("SELECT finished_at, start_date, end_date, days, rows, status, message "
                        "FROM import_runs ORDER BY finished_at DESC LIMIT 50")
            rows = cur.fetchall()
        return {"enabled": True, "runs": [
            {"finished_at": str(r[0]), "start_date": str(r[1]), "end_date": str(r[2]),
             "days": r[3], "rows": r[4], "status": r[5], "message": r[6]} for r in rows]}
    except Exception as e:  # noqa: BLE001
        return {"enabled": True, "error": str(e)[:160], "runs": []}


# ---------------------------------------------------------------------------
# External read/write API (JSON, X-API-Key auth). For other systems to manage
# mail accounts programmatically. Passwords are accepted on write, never returned.
# ---------------------------------------------------------------------------
@app.get("/api/v1/accounts", dependencies=[Depends(require_api_key)])
def api_list(active: bool | None = None):
    with SessionLocal() as s:
        q = s.query(MailAccount)
        if active is not None:
            q = q.filter(MailAccount.is_active.is_(active))
        return [_account_out(a) for a in q.order_by(MailAccount.username).all()]


@app.get("/api/v1/accounts/{username}", dependencies=[Depends(require_api_key)])
def api_get(username: str):
    with SessionLocal() as s:
        a = s.query(MailAccount).filter_by(username=username).first()
        if not a:
            raise HTTPException(404, "not found")
        return _account_out(a)


@app.post("/api/v1/accounts", status_code=201, dependencies=[Depends(require_api_key)])
def api_create(body: AccountIn):
    username = body.username.strip()
    email = body.email.strip() or f"{username}@{DEFAULT_DOMAIN}"
    kind = body.kind if body.kind in ("user", "shared") else "shared"
    with SessionLocal() as s:
        if s.query(MailAccount).filter_by(username=username).first():
            raise HTTPException(409, "username already exists")
        a = MailAccount(username=username, password=body.password or _gen_password(),
                        email=email, quota_gb=body.quota_gb, is_active=body.is_active,
                        kind=kind, notes=body.notes)
        s.add(a)
        s.commit()
        return _account_out(a)


@app.get("/api/v1/ldap/users", dependencies=[Depends(require_api_key)])
def api_ldap_users():
    """AD users (sAMAccountName) available to provision as User mailboxes."""
    return ldap_client.list_users()


@app.patch("/api/v1/accounts/{username}", dependencies=[Depends(require_api_key)])
def api_update(username: str, body: AccountPatch):
    with SessionLocal() as s:
        a = s.query(MailAccount).filter_by(username=username).first()
        if not a:
            raise HTTPException(404, "not found")
        for field, value in body.model_dump(exclude_none=True).items():
            setattr(a, field, value.strip() if isinstance(value, str) else value)
        s.commit()
        return _account_out(a)


@app.delete("/api/v1/accounts/{username}", status_code=204, dependencies=[Depends(require_api_key)])
def api_delete(username: str):
    with SessionLocal() as s:
        a = s.query(MailAccount).filter_by(username=username).first()
        if a:
            s.delete(a)
            s.commit()
