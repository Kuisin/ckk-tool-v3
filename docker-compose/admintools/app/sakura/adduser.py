#!/usr/bin/env python3
from __future__ import annotations

"""
Add users to Sakura rental server control panel from email-list.xlsx.
Loads username, password, and email per row; creates each user starting with username.
For existing users, always updates password (and mail quota) from the list so all users stay in sync.
Control panel: https://secure.sakura.ad.jp/rs/cp/users/list
"""

import argparse
import re
import sys
import time
from pathlib import Path
from typing import NamedTuple
from urllib.parse import urljoin

# Optional: use openpyxl for .xlsx (no pandas required)
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sync_playwright = None


USER_LIST_URL = "https://secure.sakura.ad.jp/rs/cp/users/list"
# Show 300 users per page to simplify edit (set via UI, not URL)
LIST_PAGE_SIZE = 300

# Usernames to skip when removing users not on list (cannot be deleted or update manually)
REMOVE_IGNORE_USERNAMES = ["postmaster"]


def _set_user_list_page_size(page, size: int = LIST_PAGE_SIZE) -> None:
    """Set users-per-page via the page-limit-selector dropdown (20/50/100/200/300件).
    Uses the right-side dropdown only; does not reload the page."""
    try:
        sel = page.locator("select.page-limit-selector").first
        if sel.count() and sel.is_visible(timeout=5000):
            sel.select_option(value=str(size))
            time.sleep(1)
            _wait_for_list_ready(page)
    except Exception:
        pass


def _is_on_user_list_page(page, list_url: str) -> bool:
    """True if current URL is the user list (and not the /new overlay path)."""
    current = page.url
    if "/new" in current:
        return False
    base = list_url.rstrip("/")
    path = current.split("?")[0].rstrip("/")
    return path == base or path.startswith(base + "/")


def _close_overlay_and_continue(page) -> None:
    """Try to close the overlay by clicking キャンセル so the script can continue."""
    try:
        page.locator(".page-overlay button:has-text('キャンセル'), .page-overlay-container button:has-text('キャンセル')").first.click()
        page.locator(".page-overlay-container").wait_for(state="hidden", timeout=8000)
    except Exception:
        pass
    time.sleep(2)


def _goto_list_page(page, list_url: str) -> None:
    """Navigate to the user list page (no /new modal). If current URL has /new, strip it to get list/1."""
    current = page.url
    if "/new" in current:
        target = current.split("/new")[0].rstrip("/")
        page.goto(target, wait_until="domcontentloaded")
    else:
        page.goto(list_url, wait_until="domcontentloaded")
    time.sleep(2)


def _collect_usernames_from_current_page(page) -> set[str]:
    """Extract existing usernames (local part before @) from .entities-body .username on current page."""
    usernames: set[str] = set()
    try:
        for s in page.locator(".entities-body .username").all_text_contents():
            t = s.strip()
            if "@" in t:
                t = t.split("@", 1)[0]
            if t:
                usernames.add(t)
    except Exception:
        pass
    return usernames


def _wait_for_list_ready(page) -> None:
    """Wait for list body to be visible after full page load (goto uses wait_until='load')."""
    try:
        page.locator(".entities-body").wait_for(state="visible", timeout=15000)
    except Exception:
        pass
    time.sleep(1)


def _find_row_with_username(page, username: str):
    """Return a locator for .entities-item row whose .username contains username (or username@...), or None."""
    try:
        items = page.locator(".entities-body .entities-item")
        for i in range(items.count()):
            row = items.nth(i)
            text = row.locator(".username").first.text_content() or ""
            t = text.strip()
            if "@" in t:
                t = t.split("@", 1)[0]
            if t.strip() == username:
                return row
    except Exception:
        pass
    return None


def _goto_page_with_username(page, list_url: str, username: str) -> bool:
    """Navigate through list pages until we find a row with this username. Return True if found."""
    current = page.url
    if "/new" in current:
        current = current.split("/new")[0].rstrip("/")
    if current != list_url and "/list" not in current:
        page.goto(list_url, wait_until="load", timeout=30000)
        _wait_for_list_ready(page)
        current = page.url
    seen_urls: set[str] = set()
    while current and current not in seen_urls:
        seen_urls.add(current)
        if _find_row_with_username(page, username) is not None:
            return True
        next_url = None
        try:
            next_btn = page.locator('a:has-text("次へ"), button:has-text("次へ"), a[rel="next"]').first
            if next_btn.count() and next_btn.is_visible(timeout=1000):
                href = next_btn.get_attribute("href")
                if href:
                    next_url = urljoin(current, href)
        except Exception:
            pass
        if not next_url or next_url == current:
            match = re.search(r"(/.+/list/)(\d+)(?:/)?", current.split("?")[0])
            if match:
                base, num = match.group(1), int(match.group(2))
                next_url = f"{base}{num + 1}"
            else:
                match = re.search(r"(/.+/list)(/)?$", current.split("?")[0])
                if match:
                    next_url = f"{match.group(1)}/2"
            if next_url and not next_url.startswith("http"):
                next_url = urljoin(current, next_url)
        if not next_url or next_url == current:
            break
        try:
            page.goto(next_url, wait_until="load", timeout=30000)
            _wait_for_list_ready(page)
            current = page.url
        except Exception:
            break
    return False


def _update_existing_user_password(
    page, user, list_url: str, mail_gb: int = 5
) -> bool:
    """Open edit form for existing user, set password (and mail quota), submit 更新. Return True if success."""
    if not _goto_page_with_username(page, list_url, user.username):
        return False
    row = _find_row_with_username(page, user.username)
    if row is None:
        return False
    # Open edit: common patterns — username link, or 編集 link/button
    edit_selectors = [
        'a[href*="edit"]',
        'a[href*="/users/"]',
        'button:has-text("編集")',
        'a:has-text("編集")',
        ".username a",
        ".entity-field a",
    ]
    edit_clicked = False
    for sel in edit_selectors:
        try:
            el = row.locator(sel).first
            if el.count() and el.is_visible(timeout=1500):
                el.click()
                edit_clicked = True
                break
        except Exception:
            continue
    if not edit_clicked:
        try:
            row.locator(".username").first.click()
            edit_clicked = True
        except Exception:
            pass
    if not edit_clicked:
        return False
    page.wait_for_load_state("domcontentloaded")
    time.sleep(1)
    form = page.locator("form.user-edit-form")
    try:
        form.wait_for(state="visible", timeout=10000)
    except Exception:
        return False
    # Fill password only (name is usually read-only on edit)
    if user.password:
        pw_el = page.locator('form.user-edit-form input[name="password"]')
        if pw_el.count():
            pw_el.fill(user.password)
    # Optionally sync mail quota
    mail_quota = page.locator('form.user-edit-form input[name="mailQuota"]')
    if mail_quota.count():
        mail_quota.fill(str(mail_gb))
        su = page.locator('form.user-edit-form select[name="mailQuotaUnit"]')
        if su.count():
            su.select_option(value="1073741824")
    # Submit: 更新 (update) for edit, not 作成する
    update_btn = page.locator(
        'form.user-edit-form button[type="submit"]:has-text("更新"), '
        'form.user-edit-form button[type="submit"]:has-text("保存"), '
        'form.user-edit-form button[type="submit"]'
    ).first
    if not update_btn.count() or not update_btn.is_visible(timeout=3000):
        try:
            page.locator(".page-overlay button:has-text('キャンセル'), .page-overlay-container button:has-text('キャンセル')").first.click()
        except Exception:
            pass
        return False
    update_btn.click()
    time.sleep(2)
    try:
        page.locator(".page-overlay-container").wait_for(state="hidden", timeout=15000)
    except Exception:
        pass
    time.sleep(1)
    return True


def _update_existing_user_password_via_settings(
    page, user, mail_gb: int = 5
) -> bool:
    """Open パスワード設定 from 設定 on the user's row (current page only, no reload).
    Fill current password (if present), new password; submit. If UI says password is unchanged, close modal and skip."""
    row = _find_row_with_username(page, user.username)
    if row is None:
        return False
    try:
        settings_btn = row.locator('button:has-text("設定")').first
        if not settings_btn.count() or not settings_btn.is_visible(timeout=2000):
            return False
        settings_btn.click()
        time.sleep(0.5)
    except Exception:
        return False
    try:
        password_setting = page.locator(
            'a:has-text("パスワード設定"), button:has-text("パスワード設定"), '
            '[role="menuitem"]:has-text("パスワード設定"), .bp5-menu-item:has-text("パスワード設定")'
        ).first
        password_setting.wait_for(state="visible", timeout=5000)
        password_setting.click()
        time.sleep(1)
    except Exception:
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return False
    page.wait_for_load_state("domcontentloaded")
    time.sleep(1.5)
    form = page.locator(
        ".page-overlay-container form.user-edit-form, .page-overlay form, "
        ".page-overlay-container form, form:has(input[type='password'])"
    ).first
    try:
        form.wait_for(state="visible", timeout=12000)
    except Exception:
        try:
            _close_overlay_and_continue(page)
        except Exception:
            pass
        return False
    if not user.password:
        try:
            _close_overlay_and_continue(page)
        except Exception:
            pass
        return False
    # Current password (1st field) and new password (2nd field); use user.password for both (if same, UI will show "same password" and we close)
    pw_inputs = form.locator('input[type="password"]')
    n = pw_inputs.count()
    if n >= 2:
        pw_inputs.nth(0).fill(user.password)
        time.sleep(0.2)
        pw_inputs.nth(1).fill(user.password)
    elif n == 1:
        pw_inputs.first.fill(user.password)
    else:
        form.locator('input[name="password"]').first.fill(user.password)
    time.sleep(0.3)
    mail_quota = form.locator('input[name="mailQuota"]')
    if mail_quota.count() and mail_quota.is_visible(timeout=500):
        mail_quota.fill(str(mail_gb))
        su = form.locator('select[name="mailQuotaUnit"]')
        if su.count():
            su.select_option(value="1073741824")
    update_btn = form.locator(
        'button[type="submit"]:has-text("更新"), '
        'button[type="submit"]:has-text("保存"), '
        'button[type="submit"]'
    ).first
    if not update_btn.count() or not update_btn.is_visible(timeout=3000):
        try:
            _close_overlay_and_continue(page)
        except Exception:
            pass
        return False
    update_btn.click()
    time.sleep(2)
    # If UI says password unchanged / same, or "past password" reuse error, close modal and skip
    skip_phrases = (
        "同じパスワード", "同一のパスワード", "変更されていません", "変更ありません", "同じです", "同一です",
        "過去に設定された事のあるパスワード", "新しいパスワードは異なる文字列で指定",
    )
    for _ in range(5):
        try:
            overlay = page.locator(".page-overlay-container")
            if overlay.count():
                overlay_text = overlay.text_content() or ""
                if any(p in overlay_text for p in skip_phrases):
                    page.locator(".page-overlay-container button:has-text('キャンセル'), .page-overlay-container button:has-text('閉じる')").first.click()
                    time.sleep(1)
                    page.locator(".page-overlay-container").wait_for(state="hidden", timeout=5000)
                    return True
        except Exception:
            pass
        page_text = page.content()
        if any(p in page_text for p in skip_phrases):
            try:
                page.locator(".page-overlay-container button:has-text('キャンセル'), .page-overlay-container button:has-text('閉じる')").first.click()
                time.sleep(1)
                page.locator(".page-overlay-container").wait_for(state="hidden", timeout=5000)
            except Exception:
                pass
            return True
        time.sleep(0.6)
    try:
        page.locator(".page-overlay-container").wait_for(state="hidden", timeout=15000)
    except Exception:
        pass
    time.sleep(1)
    return True


def _collect_existing_usernames_from_all_pages(page, list_url: str) -> set[str]:
    """Navigate through all list pages and return the set of all existing usernames."""
    all_usernames: set[str] = set()
    seen_urls: set[str] = set()
    # Start at list (no /new)
    current = page.url
    if "/new" in current:
        current = current.split("/new")[0].rstrip("/")
    if current != list_url and "/list" in current:
        page.goto(current, wait_until="load", timeout=30000)
        _wait_for_list_ready(page)
    else:
        page.goto(list_url, wait_until="load", timeout=30000)
        _wait_for_list_ready(page)
    current = page.url

    while current and current not in seen_urls:
        seen_urls.add(current)
        usernames = _collect_usernames_from_current_page(page)
        if not usernames and len(all_usernames) > 0:
            break
        all_usernames |= usernames

        # Next page: try "次へ" link, or same path with incremented page number
        next_url: str | None = None
        try:
            next_btn = page.locator('a:has-text("次へ"), button:has-text("次へ"), a[rel="next"]').first
            if next_btn.count() and next_btn.is_visible(timeout=1000):
                href = next_btn.get_attribute("href")
                if href:
                    next_url = urljoin(current, href)
        except Exception:
            pass
        if not next_url or next_url == current:
            # Try /list/1 -> /list/2, etc.
            current_path = current.split("?")[0]
            match = re.search(r"(/.+/list/)(\d+)(?:/)?$", current_path)
            if match:
                base, num = match.group(1), int(match.group(2))
                next_url = f"{base}{num + 1}"
            else:
                match = re.search(r"(/.+/list)(/)?$", current_path)
                if match:
                    next_url = f"{match.group(1)}/2"
            if next_url and not next_url.startswith("http"):
                next_url = urljoin(current, next_url)
        if not next_url or next_url == current:
            break
        try:
            page.goto(next_url, wait_until="load", timeout=30000)
            _wait_for_list_ready(page)
            current = page.url
        except Exception:
            break

    return all_usernames


def _get_next_list_page_url(page, current: str, list_url: str) -> str | None:
    """Return URL of next list page, or None if no next page."""
    next_url = None
    try:
        next_btn = page.locator('a:has-text("次へ"), button:has-text("次へ"), a[rel="next"]').first
        if next_btn.count() and next_btn.is_visible(timeout=1000):
            href = next_btn.get_attribute("href")
            if href:
                next_url = urljoin(current, href)
    except Exception:
        pass
    if not next_url or next_url == current:
        current_path = current.split("?")[0]
        match = re.search(r"(/.+/list/)(\d+)(?:/)?$", current_path)
        if match:
            base, num = match.group(1), int(match.group(2))
            next_url = f"{base}{num + 1}"
        else:
            match = re.search(r"(/.+/list)(/)?$", current_path)
            if match:
                next_url = f"{match.group(1)}/2"
            if next_url and not next_url.startswith("http"):
                next_url = urljoin(current, next_url)
        if next_url == current:
            next_url = None
    return next_url


def remove_users_not_on_list_on_page(
    page,
    keep_usernames: set[str],
    list_url: str = USER_LIST_URL,
) -> None:
    """
    Delete every user on the server that is not in keep_usernames.
    keep_usernames = usernames that are in the Excel list and active (column E true).
    Opens 設定 → 詳細設定 on each row, then clicks 削除 in the modal. Paginates through all list pages.
    After 削除 click the browser shows a native confirm (e.g. "a.haradaを削除してよろしいでしょうか。") — we accept it via the dialog listener.
    """
    def accept_dialog(dialog):
        """Accept native confirm (xxxを削除してよろしいでしょうか。) so user is deleted."""
        dialog.accept()

    page.on("dialog", accept_dialog)
    removed_count = 0
    seen_urls: set[str] = set()
    current = page.url
    if "/new" in current:
        current = current.split("/new")[0].rstrip("/")
    if current != list_url and "/list" not in current:
        page.goto(list_url, wait_until="load", timeout=30000)
        _wait_for_list_ready(page)
        current = page.url

    try:
        while current and current not in seen_urls:
            seen_urls.add(current)
            _goto_list_page(page, list_url)
            page.goto(current, wait_until="load", timeout=30000)
            _wait_for_list_ready(page)
            _set_user_list_page_size(page, LIST_PAGE_SIZE)
            time.sleep(0.5)

            deleted_one = False
            while True:
                items = page.locator(".entities-body .entities-item")
                n = items.count()
                for i in range(n):
                    try:
                        row = items.nth(i)
                        uname_el = row.locator(".username").first
                        if not uname_el.count():
                            continue
                        raw = (uname_el.text_content() or "").strip()
                        username = raw.split("@", 1)[0].strip() if "@" in raw else raw
                        if not username or username in keep_usernames or username.lower() in {u.lower() for u in REMOVE_IGNORE_USERNAMES}:
                            continue
                        settings_btn = row.locator('button:has-text("設定")').first
                        if not settings_btn.count() or not settings_btn.is_visible(timeout=1500):
                            continue
                        settings_btn.click()
                        time.sleep(0.5)
                        detail_link = page.locator(
                            'a:has-text("詳細設定"), button:has-text("詳細設定"), '
                            '[role="menuitem"]:has-text("詳細設定"), .bp5-menu-item:has-text("詳細設定")'
                        ).first
                        if not detail_link.count() or not detail_link.is_visible(timeout=2000):
                            try:
                                page.keyboard.press("Escape")
                            except Exception:
                                pass
                            continue
                        detail_link.click()
                        time.sleep(0.8)
                        page.wait_for_load_state("domcontentloaded", timeout=5000)
                        delete_btn = page.locator(
                            ".page-overlay-container button:has-text('削除'), .page-overlay button:has-text('削除'), "
                            ".page-overlay-container a:has-text('削除'), .page-overlay a:has-text('削除')"
                        ).first
                        if not delete_btn.count() or not delete_btn.is_visible(timeout=3000):
                            try:
                                _close_overlay_and_continue(page)
                            except Exception:
                                pass
                            continue
                        delete_btn.click()
                        time.sleep(1.0)  # allow native confirm "xxxを削除してよろしいでしょうか。" to appear and be accepted by listener
                        try:
                            confirm_btn = page.locator(
                                ".page-overlay button:has-text('削除'), .page-overlay-container button:has-text('削除')"
                            ).first
                            if confirm_btn.count() and confirm_btn.is_visible(timeout=1500):
                                confirm_btn.click()
                                time.sleep(0.5)
                        except Exception:
                            pass
                        try:
                            page.locator(".page-overlay-container").wait_for(state="hidden", timeout=8000)
                        except Exception:
                            pass
                        time.sleep(0.5)
                        removed_count += 1
                        print(f"  Removed user (not on list or inactive): {username}")
                        deleted_one = True
                        break
                    except Exception:
                        try:
                            page.keyboard.press("Escape")
                            _close_overlay_and_continue(page)
                        except Exception:
                            pass
                        continue
                if not deleted_one:
                    break
                current = page.url
                if "/new" in current:
                    current = current.split("/new")[0].rstrip("/")
                time.sleep(0.5)  # allow list to update in DOM without full reload

            next_url = _get_next_list_page_url(page, current, list_url)
            if not next_url or next_url == current:
                break
            try:
                page.goto(next_url, wait_until="load", timeout=30000)
                _wait_for_list_ready(page)
                _set_user_list_page_size(page, LIST_PAGE_SIZE)
                current = page.url
            except Exception:
                break
    finally:
        try:
            page.remove_listener("dialog", accept_dialog)
        except Exception:
            pass

    if removed_count:
        print(f"Removed {removed_count} user(s) that were not on the list or not active.")


class UserRow(NamedTuple):
    username: str
    password: str
    email: str


def _find_column_index(header: list[str], *candidates: str) -> int:
    """Return column index for first header that matches any candidate (substring or exact, case-insensitive)."""
    h_lower = [(h or "").lower() for h in header]
    for c in candidates:
        for i, h in enumerate(h_lower):
            if c in h or h == c:
                return i
    return -1


def _is_true(val) -> bool:
    """Treat Excel boolean/string as true if True, 'true', 'yes', '1', 'on' (case-insensitive)."""
    if val is True:
        return True
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("true", "yes", "1", "on")


def load_users_from_xlsx(path: str | Path, sheet_name: str | None = None) -> list[UserRow]:
    """Load users from Excel: username, password, email (with domain).
    sheet_name: if given, use that sheet; else use 'other' if present, else active.
    Only includes rows where column E is true. Rows without username are skipped."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Email list not found: {path}")

    if openpyxl is None:
        raise ImportError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb["other"] if "other" in wb.sheetnames else wb.active
    if ws is None:
        wb.close()
        return []

    users: list[UserRow] = []
    header = None
    username_idx = -1
    password_idx = -1
    email_idx = -1

    for row in ws.iter_rows(values_only=True):
        row = [str(c).strip() if c is not None else "" for c in (row or [])]
        if not row:
            continue
        if header is None:
            header = row
            username_idx = _find_column_index(
                header, "username", "ユーザー名", "user", "ログイン名", "login"
            )
            password_idx = _find_column_index(
                header, "password", "パスワード", "pass"
            )
            # Prefer full email column
            for i, h in enumerate(header):
                h_lower = (h or "").lower()
                if "domain" in h_lower and ("email" in h_lower or "mail" in h_lower or "メール" in h_lower):
                    email_idx = i
                    break
            else:
                email_idx = _find_column_index(
                    header, "email (with domain)", "email", "メール", "mail", "e-mail", "メールアドレス"
                )
            continue
        if username_idx < 0:
            continue
        # Column E (index 4): only create user when true
        if not _is_true(row[4] if len(row) > 4 else None):
            continue
        username = row[username_idx].strip() if username_idx < len(row) else ""
        if not username:
            continue
        password = row[password_idx].strip() if 0 <= password_idx < len(row) else ""
        email = row[email_idx].strip() if 0 <= email_idx < len(row) else ""
        if not email and email_idx >= 0 and email_idx < len(row):
            email = row[email_idx].strip()
        users.append(UserRow(username=username, password=password, email=email))
    wb.close()
    return users


def add_users_on_page(
    page,
    users: list[UserRow],
    url: str = USER_LIST_URL,
    pause_before_add: bool = False,
    start_from: int = 1,
    total_count: int | None = None,
    mail_gb: int = 5,
    update_existing_password: bool = True,
    no_reload: bool = False,
) -> None:
    """Add users on an already-open, logged-in page (user list). For existing users, updates password if update_existing_password.
    When no_reload=True: do not goto or set page size; use current page only; open パスワード設定 from 設定 per user."""
    total = total_count if total_count is not None else len(users)
    page.locator(".page-overlay-container").wait_for(state="hidden", timeout=5000)
    if no_reload:
        existing_usernames = _collect_usernames_from_current_page(page)
        print(f"Existing users on current page: {len(existing_usernames)}.")
    else:
        if _is_on_user_list_page(page, url):
            _set_user_list_page_size(page, LIST_PAGE_SIZE)
        else:
            _goto_list_page(page, url)
            _set_user_list_page_size(page, LIST_PAGE_SIZE)
        print("Checking all pages for existing users...")
        existing_usernames = _collect_existing_usernames_from_all_pages(page, url)
        _goto_list_page(page, url)
        _set_user_list_page_size(page, LIST_PAGE_SIZE)

    already_in_list = [u for u in users if u.username in existing_usernames]
    to_add_count = len(users) - len(already_in_list)
    if not no_reload:
        print(f"\nExisting users (all pages): {len(existing_usernames)}.")
    if already_in_list:
        print(f"From your list, {len(already_in_list)} already exist (will update password): {', '.join(u.username for u in already_in_list)}")
    print(f"Will add: {to_add_count} new user(s). Will update password for existing.")
    if to_add_count == 0 and not already_in_list:
        print("Nothing to do.")
        return

    for j, user in enumerate(users, 1):
            i = start_from + j - 1
            print(f"\n[{i}/{total}] Adding user: {user.username}" + (f" ({user.email})" if user.email else ""))
            if pause_before_add and i > 1:
                input("Press Enter to continue to next user...")

            try:
                # 2人目以降: オーバーレイが閉じるまで待ち、一覧が安定するまで少し待つ
                if j > 1:
                    try:
                        page.locator(".page-overlay-container").wait_for(state="hidden", timeout=15000)
                    except Exception:
                        _close_overlay_and_continue(page)
                    time.sleep(2)

                # 既存ユーザー: パスワードを常に更新（update_existing_password が True の場合）
                if user.username in existing_usernames:
                    if update_existing_password and user.password:
                        print(f"  Updating password for existing user: {user.username}")
                        ok = (
                            _update_existing_user_password_via_settings(page, user, mail_gb)
                            if no_reload
                            else _update_existing_user_password(page, user, url, mail_gb)
                        )
                        if ok:
                            print(f"  Updated password for {user.username}.")
                        else:
                            print(f"  Could not open パスワード設定 for {user.username}; update password manually.")
                    else:
                        print(f"  Skip: {user.username} already exists." + (" (no password in list)" if not user.password else ""))
                    continue
                try:
                    existing_texts = page.locator(".entities-body .username").all_text_contents()
                    prefix = user.username + "@"
                    if any(s.strip().startswith(prefix) for s in existing_texts):
                        existing_usernames.add(user.username)
                        if update_existing_password and user.password:
                            print(f"  Updating password for existing user: {user.username}")
                            ok = (
                                _update_existing_user_password_via_settings(page, user, mail_gb)
                                if no_reload
                                else _update_existing_user_password(page, user, url, mail_gb)
                            )
                            if ok:
                                print(f"  Updated password for {user.username}.")
                            else:
                                print(f"  Could not open パスワード設定 for {user.username}; update password manually.")
                        else:
                            print(f"  Skip: {user.username} already exists." + (" (no password in list)" if not user.password else ""))
                        continue
                except Exception:
                    pass  # 一覧がまだない等は無視して追加を試行

                # 「新規追加」ボタンを待ってクリック（オーバーレイが閉じきるまでリトライ）
                add_selectors = [
                    "[data-testid='page-content-global-action'] a",
                    "a[href*='/users/list'][href*='/new']",
                    "a:has-text('新規追加')",
                    "button:has-text('新規追加')",
                    "input[value='新規追加']",
                ]
                add_btn = None
                for attempt in range(3):
                    if attempt > 0:
                        page.locator(".page-overlay-container").wait_for(state="hidden", timeout=10000)
                        time.sleep(2)
                    for sel in add_selectors:
                        try:
                            el = page.locator(sel).first
                            if el.count() and el.is_visible(timeout=5000):
                                add_btn = el
                                break
                        except Exception:
                            continue
                    if add_btn is not None:
                        break
                if add_btn is None:
                    print("  Could not find 'Add user' button. Please add this user manually.")
                    continue
                add_btn.click()
                page.wait_for_load_state("domcontentloaded")
                time.sleep(0.5)

                # さくらパネル: オーバーレイ内の form.user-edit-form を待つ
                form = page.locator("form.user-edit-form")
                form.wait_for(state="visible", timeout=10000)

                # 1) ユーザー名 (メールの@左側 = input[name="name"])
                page.locator('form.user-edit-form input[name="name"]').fill(user.username)

                # 2) パスワード (必須)
                if user.password:
                    page.locator('form.user-edit-form input[name="password"]').fill(user.password)

                # 3) メール利用: 「利用する」にチェック
                uses_mail = page.locator('form.user-edit-form input[name="usesMail"]')
                if uses_mail.count() and not uses_mail.is_checked():
                    uses_mail.check()

                # 4) メール容量制限 (必須項目)
                mail_quota = page.locator('form.user-edit-form input[name="mailQuota"]')
                if mail_quota.count():
                    mail_quota.fill(str(mail_gb))
                    page.locator('form.user-edit-form select[name="mailQuotaUnit"]').select_option(value="1073741824")  # GB

                # 5) メールの受信: 「受信する」を選択
                page.locator('form.user-edit-form label:has-text("受信する")').locator('input[name="mailReceiveType"]').check()

                # 6) 作成する
                page.locator('form.user-edit-form button[type="submit"]:has-text("作成する")').click()
                time.sleep(2)  # 応答（成功 or 重複エラー）を待つ

                # 重複エラー「すでに存在しています」をチェック（数秒間ポーリング）
                is_duplicate = False
                for _ in range(6):
                    try:
                        if page.get_by_text("すでに存在しています").is_visible(timeout=1500):
                            is_duplicate = True
                            break
                    except Exception:
                        pass
                    time.sleep(0.5)
                if is_duplicate:
                    existing_usernames.add(user.username)
                    print(f"  Skip: {user.username} already exists (duplicate)." + ("" if no_reload else " Going to list page."))
                    if not no_reload:
                        _goto_list_page(page, url)
                    continue

                # 成功: オーバーレイが閉じるまで待つ。閉じない場合はリロードして次へ（no_reload のときはリロードしない）
                try:
                    page.locator(".page-overlay-container").wait_for(state="hidden", timeout=25000)
                    existing_usernames.add(user.username)
                    time.sleep(2)
                    print(f"  Submitted for {user.username}. Verify in the browser.")
                except Exception:
                    existing_usernames.add(user.username)
                    print(f"  Overlay did not close in time." + ("" if no_reload else " Going to list page."))
                    if no_reload:
                        _close_overlay_and_continue(page)
                    else:
                        _goto_list_page(page, url)
            except Exception as e:
                print(f"  Error: {e}. Please add {user.username} manually.")

    print("\nAdd-users step done.")


def add_users_with_browser(
    users: list[UserRow],
    url: str = USER_LIST_URL,
    headless: bool = False,
    pause_before_add: bool = False,
    start_from: int = 1,
    total_count: int | None = None,
    mail_gb: int = 5,
    update_existing_password: bool = True,
    no_reload: bool = False,
) -> None:
    total = total_count if total_count is not None else len(users)
    if sync_playwright is None:
        raise ImportError("Install Playwright: pip install playwright && playwright install chromium")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(
            locale="ja-JP",
            viewport={"width": 1280, "height": 900},
        )
        page = context.new_page()

        print(f"Opening {url}")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)

        print("\n--- Please log in to the control panel in the browser. ---")
        input("Press Enter here after you have logged in and the user list page is visible...")

        add_users_on_page(
            page, users, url=url,
            pause_before_add=pause_before_add,
            start_from=start_from,
            total_count=total_count,
            mail_gb=mail_gb,
            update_existing_password=update_existing_password,
            no_reload=no_reload,
        )

        print("\nDone. You can close the browser or press Enter to close.")
        input()
        browser.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Add users to Sakura control panel from an Excel list (username, password, email)."
    )
    parser.add_argument(
        "excel_file",
        nargs="?",
        default=Path(__file__).parent / "email-list.xlsx",
        help="Path to email-list.xlsx (default: email-list.xlsx in script directory)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only load and print emails, do not open browser",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run browser in headless mode (not recommended for first-time login)",
    )
    parser.add_argument(
        "--pause",
        action="store_true",
        help="Pause before each user add (after the first)",
    )
    parser.add_argument(
        "--start-from",
        type=int,
        default=1,
        metavar="N",
        help="Start from user N (1-based). E.g. --start-from 52 to resume from 52nd user.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        metavar="NAME",
        help="Sheet name to read from (default: 'other' if present, else first/active sheet).",
    )
    parser.add_argument(
        "--mail-gb",
        type=int,
        default=5,
        metavar="N",
        help="Mail inbox quota in GB (default: 5).",
    )
    parser.add_argument(
        "--no-update-password",
        action="store_true",
        help="Do not update password for existing users (default: always update password from list).",
    )
    parser.add_argument(
        "--no-reload",
        action="store_true",
        help="Do not reload page or set list size; use current page only. Set list to 300 manually before starting.",
    )
    args = parser.parse_args()

    try:
        all_users = load_users_from_xlsx(args.excel_file, sheet_name=args.sheet)
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        return 1
    except ValueError as e:
        print(e, file=sys.stderr)
        return 1
    except ImportError as e:
        print(e, file=sys.stderr)
        print("Run: pip install openpyxl playwright && playwright install chromium", file=sys.stderr)
        return 1

    if not all_users:
        print("No users found (need username column) in the Excel file.", file=sys.stderr)
        return 1

    start_from = max(1, min(args.start_from, len(all_users)))
    users = all_users[start_from - 1:]
    if not users:
        print("No users to process (--start-from is beyond list).", file=sys.stderr)
        return 0

    print(f"Loaded {len(all_users)} user(s) total." + (f" Starting from #{start_from} ({len(users)} remaining)." if start_from > 1 else ""))
    for u in users:
        print(f"  - {u.username} | {'***' if u.password else '-'} | {u.email or '-'}")

    if args.dry_run:
        print("Dry run: not opening browser.")
        return 0

    try:
        add_users_with_browser(
            users,
            url=USER_LIST_URL,
            headless=args.headless,
            pause_before_add=args.pause,
            start_from=start_from,
            total_count=len(all_users),
            mail_gb=args.mail_gb,
            update_existing_password=not args.no_update_password,
            no_reload=args.no_reload,
        )
    except ImportError as e:
        print(e, file=sys.stderr)
        print("Run: pip install playwright && playwright install chromium", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
