#!/usr/bin/env python3
from __future__ import annotations

"""
Create mail aliases (email addresses) for users where the email local part
differs from the username. Uses メール エイリアス (mail alias) form.
Reads from email-list.xlsx; only processes rows where column E is active and email ≠ username.

Before adding: removes any alias on the server that is not on the list or not active
(so the server stays in sync with the Excel list). Use --no-remove to skip removal.

Control panel: https://secure.sakura.ad.jp/rs/cp/users/mail-alias
"""

import argparse
import sys
import time
from pathlib import Path
from typing import NamedTuple

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sync_playwright = None


MAIL_ALIAS_LIST_URL = "https://secure.sakura.ad.jp/rs/cp/users/mail-alias"
MAIL_ALIAS_NEW_URL = "https://secure.sakura.ad.jp/rs/cp/users/mail-alias/new"


def _goto_list_page(page, list_url: str) -> None:
    """Navigate to the alias list page (no /new). If current URL has /new, strip it."""
    current = page.url
    if "/new" in current:
        target = current.split("/new")[0].rstrip("/")
        page.goto(target, wait_until="load", timeout=30000)
    else:
        page.goto(list_url, wait_until="load", timeout=30000)
    time.sleep(2)


def _collect_addresses_from_current_page(page) -> set[str]:
    """Extract existing alias addresses (local@domain) from .entities-body .col-address."""
    addresses: set[str] = set()
    try:
        for s in page.locator(".entities-body .col-address").all_text_contents():
            t = (s or "").strip()
            if t and "@" in t:
                addresses.add(t)
    except Exception:
        pass
    return addresses


def _wait_for_list_ready(page) -> None:
    """Wait for list body to be visible after full page load."""
    try:
        page.locator(".entities-body").wait_for(state="visible", timeout=15000)
    except Exception:
        pass
    time.sleep(1)


def _collect_existing_aliases(page, list_url: str) -> set[str]:
    """Load the mail-alias list page (single page, no pagination) and return all existing addresses (local@domain)."""
    current = page.url
    if "/new" in current:
        current = current.split("/new")[0].rstrip("/")
    if current != list_url and "mail-alias" in current:
        page.goto(current, wait_until="load", timeout=30000)
        _wait_for_list_ready(page)
    else:
        page.goto(list_url, wait_until="load", timeout=30000)
        _wait_for_list_ready(page)
    return _collect_addresses_from_current_page(page)


def _remove_aliases_not_on_list_on_page(
    page,
    keep_addresses: set[str],
    list_url: str = MAIL_ALIAS_LIST_URL,
) -> None:
    """
    Delete every mail alias on the server that is not in keep_addresses.
    keep_addresses = addresses that are on the Excel list and active (column E true, email ≠ username).
    Each row has .entities-item with .col-address and a 削除 button.
    After 削除 click the browser shows a native confirm: "このエイリアス設定を削除しますか？" — we accept it via the dialog listener.
    """
    _goto_list_page(page, list_url)
    page.locator(".entities-body").wait_for(state="visible", timeout=15000)
    time.sleep(0.5)

    def accept_dialog(dialog):
        """Accept native confirm (このエイリアス設定を削除しますか？) so alias is deleted."""
        dialog.accept()

    page.on("dialog", accept_dialog)
    removed_count = 0
    try:
        while True:
            items = page.locator(".entities-body .entities-item")
            n = items.count()
            deleted_one = False
            for i in range(n):
                try:
                    addr_el = items.nth(i).locator(".col-address").first
                    addr = (addr_el.text_content() or "").strip()
                    if not addr or "@" not in addr:
                        continue
                    if addr in keep_addresses:
                        continue
                    # This alias is not on the list or not active — delete it
                    delete_btn = items.nth(i).locator('button:has-text("削除")').first
                    if delete_btn.count() and delete_btn.is_visible(timeout=2000):
                        delete_btn.click()
                        time.sleep(1.0)  # allow native confirm "このエイリアス設定を削除しますか？" to appear and be accepted by listener
                        # If overlay confirm exists instead, click 削除 there too
                        try:
                            confirm_btn = page.locator(".page-overlay button:has-text('削除'), .page-overlay-container button:has-text('削除')").first
                            if confirm_btn.count() and confirm_btn.is_visible(timeout=2000):
                                confirm_btn.click()
                                time.sleep(0.5)
                        except Exception:
                            pass
                        try:
                            page.locator(".page-overlay-container").wait_for(state="hidden", timeout=10000)
                        except Exception:
                            pass
                        time.sleep(0.5)
                        removed_count += 1
                        print(f"  Removed alias (not on list or inactive): {addr}")
                        deleted_one = True
                        break
                except Exception:
                    continue
            if not deleted_one:
                break
            _goto_list_page(page, list_url)
            time.sleep(0.5)
    finally:
        page.remove_listener("dialog", accept_dialog)

    if removed_count:
        print(f"Removed {removed_count} alias(es) that were not on the list or not active.")


class AliasRow(NamedTuple):
    """Username (receive user) and the alias email (local part + domain)."""
    username: str
    local_part: str
    domain: str


def _find_column_index(header: list[str], *candidates: str) -> int:
    h_lower = [(h or "").lower() for h in header]
    for c in candidates:
        for i, h in enumerate(h_lower):
            if c in h or h == c:
                return i
    return -1


# Column E is index 4 (0-based). Treat as true: True, "true", "yes", "1", "○", "〇", "x"
def _is_true(val: object) -> bool:
    if val is True:
        return True
    if val is None:
        return False
    s = str(val).strip().lower()
    if s in ("", "0", "false", "no", "×", "ー", "-"):
        return False
    if s in ("1", "true", "yes", "○", "〇", "x", "✓", "✔"):
        return True
    return bool(s)


def load_alias_rows_from_xlsx(
    path: str | Path,
    sheet_name: str | None = None,
) -> list[AliasRow]:
    """
    Load from Excel; return only rows where column E is true and email local part differs from username.
    Email column should contain full address (e.g. kaisei@ckk-tool.co.jp).
    Column E (index 4) must be true/yes/1/○ etc. for the row to be included.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Email list not found: {path}")
    if openpyxl is None:
        raise ImportError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    if ws is None:
        wb.close()
        return []

    rows: list[AliasRow] = []
    header = None
    username_idx = -1
    email_idx = -1
    col_e_idx = 4  # Column E (0-based)

    for row in ws.iter_rows(values_only=True):
        row = [str(c).strip() if c is not None else "" for c in (row or [])]
        if not row:
            continue
        if header is None:
            header = row
            username_idx = _find_column_index(
                header, "username", "ユーザー名", "user", "ログイン名", "login"
            )
            for i, h in enumerate(header):
                h_lower = (h or "").lower()
                if "domain" in h_lower and ("email" in h_lower or "mail" in h_lower or "メール" in h_lower):
                    email_idx = i
                    break
            else:
                email_idx = _find_column_index(
                    header, "email (with domain)", "email", "メール", "mail", "e-mail", "メールアドレス"
                )
            continue
        if username_idx < 0 or email_idx < 0:
            continue
        # Only include rows where column E is true
        if col_e_idx >= len(row) or not _is_true(row[col_e_idx]):
            continue
        username = row[username_idx].strip() if username_idx < len(row) else ""
        email = row[email_idx].strip() if email_idx < len(row) else ""
        if not username or not email or "@" not in email:
            continue
        local_part, _, domain = email.partition("@")
        local_part = local_part.strip()
        domain = domain.strip()
        if not local_part or not domain:
            continue
        # Only add alias when email local part differs from username
        if local_part != username:
            rows.append(AliasRow(username=username, local_part=local_part, domain=domain))
    wb.close()
    return rows


def add_aliases_on_page(
    page,
    aliases: list[AliasRow],
    url: str = MAIL_ALIAS_LIST_URL,
    remove_not_on_list: bool = True,
) -> None:
    """Create mail aliases on an already-open, logged-in page (mail-alias list). Used by add_aliases_with_browser and combined script.
    If remove_not_on_list, first removes any server alias that is not on the Excel list or not active (column E). Then adds missing aliases."""
    page.locator(".page-overlay-container").wait_for(state="hidden", timeout=5000)
    print("Loading existing aliases...")
    existing_set = _collect_existing_aliases(page, url)
    _goto_list_page(page, url)

    if remove_not_on_list:
        # Remove aliases that are not on the list or not active (column E)
        keep_addresses = {f"{a.local_part}@{a.domain}" for a in aliases}
        print("Removing aliases not on list or not active...")
        _remove_aliases_not_on_list_on_page(page, keep_addresses, url)
        # Re-load existing after removals
        existing_set = _collect_existing_aliases(page, url)
        _goto_list_page(page, url)

    already_exist = [a for a in aliases if f"{a.local_part}@{a.domain}" in existing_set]
    aliases_to_create = [a for a in aliases if f"{a.local_part}@{a.domain}" not in existing_set]

    print(f"\nExisting aliases on server: {len(existing_set)}.")
    if already_exist:
        print(f"From your list, already exist ({len(already_exist)}):")
        for a in already_exist:
            print(f"  - {a.local_part}@{a.domain} → {a.username}")
    print(f"Will create: {len(aliases_to_create)} alias(es).")
    for a in aliases_to_create:
        print(f"  - {a.local_part}@{a.domain} → {a.username}")

    if not aliases_to_create:
        print("\nNothing to create.")
        return

    add_link = page.locator('a[href*="mail-alias/new"]').first
    for i, alias in enumerate(aliases_to_create, 1):
        print(f"\n[{i}/{len(aliases_to_create)}] Alias: {alias.local_part}@{alias.domain} → {alias.username}")
        try:
            if i > 1:
                add_link.wait_for(state="visible", timeout=15000)
                time.sleep(0.3)

            full_addr = f"{alias.local_part}@{alias.domain}"
            if full_addr in existing_set:
                print(f"  Skip: {full_addr} already exists.")
                continue
            try:
                existing = page.locator(".entities-body .col-address").all_text_contents()
                if any(full_addr in (t or "") for t in existing):
                    existing_set.add(full_addr)
                    print(f"  Skip: {full_addr} already exists.")
                    continue
            except Exception:
                pass

            # Click 新規追加
            for _ in range(3):
                try:
                    if add_link.is_visible(timeout=5000):
                        add_link.click()
                        break
                except Exception:
                    page.locator(".page-overlay-container").wait_for(state="hidden", timeout=10000)
                    time.sleep(2)
            else:
                print("  Could not find '新規追加' link.")
                continue

            page.wait_for_load_state("domcontentloaded")
            time.sleep(0.5)

            # Wait for overlay form (エイリアス追加)
            page.locator('input[name="localPart"]').wait_for(state="visible", timeout=10000)

            # メールアドレス (ローカル部)
            page.locator('input[name="localPart"]').fill(alias.local_part)
            # ドメイン
            page.locator('select[name="domain"]').select_option(value=alias.domain)
            # 受信先ユーザー
            page.locator('select[name="receiveUser"]').select_option(value=alias.username)

            # 追加
            page.locator('button[type="submit"]:has-text("追加")').click()
            time.sleep(1)

            # Duplicate error?
            dup = page.get_by_text("すでに存在しています")
            try:
                if dup.is_visible(timeout=2000):
                    existing_set.add(full_addr)
                    print(f"  Skip: alias already exists (duplicate).")
                    page.locator(".page-overlay button:has-text('キャンセル'), .page-overlay-container button:has-text('キャンセル')").first.click()
                    add_link.wait_for(state="visible", timeout=10000)
                    time.sleep(0.3)
                    continue
            except Exception:
                pass

            # Continue to next as soon as 新規追加 is available (no need to wait for full overlay close)
            add_link.wait_for(state="visible", timeout=20000)
            existing_set.add(full_addr)
            time.sleep(0.3)
            print(f"  Created {alias.local_part}@{alias.domain} → {alias.username}")
        except Exception as e:
            print(f"  Error: {e}. Please add alias manually.")

    print("\nAdd-aliases step done.")


def add_aliases_with_browser(
    aliases: list[AliasRow],
    url: str = MAIL_ALIAS_LIST_URL,
    headless: bool = False,
    remove_not_on_list: bool = True,
) -> None:
    if sync_playwright is None:
        raise ImportError("Install Playwright: pip install playwright && playwright install chromium")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(locale="ja-JP", viewport={"width": 1280, "height": 900})
        page = context.new_page()

        print(f"Opening {url}")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        print("\n--- Please log in to the control panel in the browser. ---")
        input("Press Enter after you have logged in and the メール エイリアス page is visible...")

        add_aliases_on_page(page, aliases, url=url, remove_not_on_list=remove_not_on_list)

        print("\nDone. Close the browser or press Enter.")
        input()
        browser.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create mail aliases (email ≠ username) from email-list.xlsx via メール エイリアス."
    )
    parser.add_argument(
        "excel_file",
        nargs="?",
        default=Path(__file__).parent / "email-list.xlsx",
        help="Path to Excel file (default: email-list.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        metavar="NAME",
        help="Sheet name to read (default: active sheet)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Only list aliases to create")
    parser.add_argument("--headless", action="store_true", help="Run browser headless")
    parser.add_argument(
        "--no-remove",
        action="store_true",
        help="Do not remove aliases that are not on the list or not active (default: remove them first)",
    )
    parser.add_argument(
        "--start-from",
        type=int,
        default=1,
        metavar="N",
        help="Start from Nth alias (1-based), e.g. 20 to skip first 19",
    )
    args = parser.parse_args()

    try:
        aliases = load_alias_rows_from_xlsx(args.excel_file, sheet_name=args.sheet)
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        return 1
    except ValueError as e:
        print(e, file=sys.stderr)
        return 1
    except ImportError as e:
        print(e, file=sys.stderr)
        return 1

    if not aliases:
        print(
            "No aliases to create (column E must be true and email local part ≠ username).",
            file=sys.stderr,
        )
        return 0

    # Apply --start-from (1-based)
    if args.start_from > 1:
        if args.start_from > len(aliases):
            print(f"start-from {args.start_from} is beyond list length {len(aliases)}.", file=sys.stderr)
            return 1
        aliases = aliases[args.start_from - 1 :]
        print(f"Starting from alias #{args.start_from} ({len(aliases)} aliases).")

    print(f"Will create up to {len(aliases)} mail alias(es) (email name ≠ username):")
    for a in aliases:
        print(f"  - {a.local_part}@{a.domain} → receive user: {a.username}")

    if args.dry_run:
        print("Dry run: not opening browser.")
        return 0

    try:
        add_aliases_with_browser(
            aliases,
            url=MAIL_ALIAS_LIST_URL,
            headless=args.headless,
            remove_not_on_list=not args.no_remove,
        )
    except ImportError as e:
        print(e, file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
