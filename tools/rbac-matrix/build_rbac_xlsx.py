#!/usr/bin/env python3
"""build_rbac_xlsx.py — _docs/rbac-role-matrix.xlsx を実データから作り直す。

前の版はセッション限りのスクリプトで生成されていたため、シードを変えても誰も
作り直せなかった（操作コードもロール構成も古いまま残っていた）。この版はリポジトリに
置き、2 つの実体だけを読む:

  1. app.roles / app.permissions / app.role_permission_relation（dev DB の実データ）
     + app.user_permission_summary（ユーザー 1 行 = 有効なロール割当）
  2. coolify/apps/nextjs-web/src/lib/app-list.ts（アプリ → 権限コードの正）

つまり出力は「シードにこう書いたつもり」ではなく「いま DB がこうなっている」。

使い方（リポジトリルートから。DB はトンネル経由で読む）:

    cd shared-db
    ./scripts/remote-db.sh python3 ../tools/rbac-matrix/build_rbac_xlsx.py

    # 直接 DATABASE_URL を指定する場合（LAN からポートが見えるとき）
    DATABASE_URL=postgresql://... python3 tools/rbac-matrix/build_rbac_xlsx.py

必要なもの: psql（PATH 上）と openpyxl。DB へは psql の CSV 出力経由で読むので、
Python 側の DB ドライバは要らない（この Mac に psycopg は入っていない）。
"""

from __future__ import annotations

import csv
import io
import os
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
APP_LIST = REPO / "coolify/apps/nextjs-web/src/lib/app-list.ts"
OUT = REPO / "_docs/rbac-role-matrix.xlsx"

# 表示順（業務フロー順）。ここに無いロールは末尾へ回す。
ROLE_ORDER = [
    "admin",
    "manager",
    "sales_manager",
    "sales",
    "sales_assistant",
    "purchasing_manager",
    "purchasing",
    "production_manager",
    "production",
    "quality_manager",
    "quality",
    "shipping_manager",
    "shipping",
    "accounting_manager",
    "accounting",
    "viewer",
    "staff",
]

# 権限コードの並び（業務フロー順）。app-list.ts の登場順とほぼ同じ。
CODE_ORDER = [
    "price_list",
    "quote",
    "order_acceptance",
    "design_request",
    "purchase_order",
    "material_receipt",
    "outsource_order",
    "work_order",
    "approve",
    "inventory",
    "delivery_order",
    "delivery_note",
    "invoice",
    "billing_closing",
    "master",
    "admin_manual",
    "kiosk",
    "system",
]

ACTION_LETTER = {
    "READ": "R",
    "CREATE": "C",
    "UPDATE": "U",
    "DELETE": "D",
    "EXPORT": "E",
    "APPROVE": "A",
    "ADMIN": "◎",
}
# 表示順（RCUDEA）。集合を並べ替えるのに使う。
LETTER_ORDER = "RCUDEA◎"

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
ADMIN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def query(sql: str) -> list[dict[str, str]]:
    """psql に CSV で吐かせて読む（DB ドライバ不要）。"""
    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("DATABASE_URL が未設定です（remote-db.sh 経由で実行してください）")
    proc = subprocess.run(
        ["psql", url, "-X", "-A", "--csv", "-c", sql],
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        sys.exit(f"psql failed: {proc.stderr.strip()}")
    return list(csv.DictReader(io.StringIO(proc.stdout)))


def read_app_list() -> list[dict[str, str]]:
    """app-list.ts から アプリ（key / label / opcode / 権限コード）を読む。"""
    src = APP_LIST.read_text(encoding="utf-8")
    pattern = re.compile(
        r'key:\s*"(?P<key>[^"]+)",\s*'
        r'label:\s*"(?P<label>[^"]+)",\s*'
        r'operationCode:\s*"(?P<code>[^"]+)",\s*'
        r'href:\s*"(?P<href>[^"]+)",\s*'
        r'icon:\s*"[^"]+",\s*'
        r'category:\s*"(?P<category>[^"]+)",\s*'
        r"(?:\s*//[^\n]*\n\s*)*"
        r"requiredPermission:\s*(?P<perm>null|\"[a-z_]+\")",
    )
    apps = []
    for m in pattern.finditer(src):
        d = m.groupdict()
        d["perm"] = "" if d["perm"] == "null" else d["perm"].strip('"')
        apps.append(d)
    if not apps:
        sys.exit("app-list.ts を解析できませんでした（書式が変わった可能性）")
    return apps


def sort_letters(letters: set[str]) -> str:
    return "".join(sorted(letters, key=LETTER_ORDER.index))


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def main() -> None:
    roles = query(
        "SELECT rolename, display_name->>'ja' AS ja, description->>'ja' AS memo, "
        "is_system FROM app.roles ORDER BY rolename"
    )
    perms = query(
        "SELECT code, display_name->>'ja' AS ja, description->>'ja' AS memo "
        "FROM app.permissions ORDER BY code"
    )
    grants = query(
        "SELECT r.rolename, rp.permission_code AS code, rp.action::text AS action, "
        "rp.scope::text AS scope, array_to_string(rp.scope_values, ',') AS scope_values "
        "FROM app.role_permission_relation rp "
        "JOIN app.roles r ON r.id = rp.role_id"
    )
    # 1 行 = 1 ユーザーのビュー（有効な割当だけ。失効日を過ぎた割当も外れる —
    # 以前の手書き GROUP BY は is_active しか見ておらず、deactivate_at を無視していた）。
    users = query(
        "SELECT username, display_name, is_active, "
        "array_to_string(roles, ',') AS roles "
        "FROM app.user_permission_summary ORDER BY username"
    )
    apps = read_app_list()

    perm_ja = {p["code"]: p["ja"] for p in perms}
    role_ja = {r["rolename"]: r["ja"] for r in roles}
    role_memo = {r["rolename"]: r["memo"] for r in roles}

    # (role, code) → { letters, scopes }
    cells: dict[tuple[str, str], dict[str, set[str]]] = {}
    for g in grants:
        key = (g["rolename"], g["code"])
        cell = cells.setdefault(key, {"letters": set(), "scopes": set()})
        cell["letters"].add(ACTION_LETTER.get(g["action"], g["action"][0]))
        cell["scopes"].add(g["scope"])

    role_names = [r for r in ROLE_ORDER if r in role_ja]
    role_names += [
        r["rolename"]
        for r in roles
        if r["rolename"] not in role_names and r["is_system"] == "t"
    ]
    codes = [c for c in CODE_ORDER if c in perm_ja]
    codes += [p["code"] for p in perms if p["code"] not in codes]

    # アプリ側から見た権限コード（マトリクスの行ラベルに操作コードを添えるため）
    code_apps: dict[str, list[str]] = {}
    for a in apps:
        if a["perm"]:
            code_apps.setdefault(a["perm"], []).append(f"{a['code']} {a['label']}")

    wb = Workbook()

    # ── はじめに ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "はじめに"
    lines = [
        ("CKK 業務管理システム — ロール・権限リファレンス", True),
        ("", False),
        ("このブックは dev データベースの実データ（app.role_permission_relation）から自動生成しています。", False),
        ("再生成: tools/rbac-matrix/build_rbac_xlsx.py", False),
        ("  cd shared-db && ./scripts/remote-db.sh python3 ../tools/rbac-matrix/build_rbac_xlsx.py", False),
        ("シードの正: shared-db/sql/rbac-seed.sql（権限コード・admin/staff）+ roles-seed.sql（運用ロール）", False),
        ("", False),
        ("記号: R=閲覧 C=作成 U=更新 D=削除 E=エクスポート ◎=ADMIN（全アクション）", True),
        ("スコープ: 無印=ALL（全件） / OWN=自分が作成した行のみ / PLANT=所属拠点 / REGION=所属地域", False),
        ("  スコープの解決は packages/authz-core の decide()。ALL 行が 1 本でもあれば無制限。", False),
        ("", False),
        ("重要: 承認は権限アクションでは管理しません（旧 A=承認 グラントは全廃）。", True),
        ("誰が承認できるかは『承認設定』（MS0B）の承認グループ所属だけで決まり、", False),
        ("権限側の要件はその書類の閲覧（R）または更新（U）を持っていることだけです。", False),
        ("", False),
        ("重要: このブックは『何ができるか』であって『何が見えるか』ではありません。", True),
        ("本番（main）のランチャー表示は feature_flags が別に決めます（shared-db/sql/feature-flags-seed.sql）。", False),
        ("権限があっても未公開のアプリは本番のホーム・ランチャーに出ません。", False),
        ("", False),
        ("本番投入手順:", True),
        ("  1. shared-db/sql/rbac-seed.sql → roles-seed.sql を本番 DB に適用（冪等）", False),
        ("  2. 実ユーザー（SSO 初回ログインで app.users に作成される）へ user_role_relation で割当", False),
        ("     割当画面: SY01 ユーザー管理 /settings/users", False),
        ("  3. 承認者は承認グループ（MS0B 承認設定）へメンバー追加 — 承認の可否はこれだけで決まる", False),
        ("  ※ dev_* / demo* ユーザーは検証専用 — 本番 DB には投入しないこと", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if i == 1 else 11)
    ws.column_dimensions["A"].width = 110

    # ── ロール一覧 ────────────────────────────────────────────────────────
    ws = wb.create_sheet("ロール一覧")
    ws.append(["ロール名 (rolename)", "表示名", "説明・運用メモ", "権限コード数", "システムロール"])
    style_header(ws, 1, 5)
    for name in role_names:
        n_codes = len({c for (r, c) in cells if r == name})
        is_sys = next((r["is_system"] for r in roles if r["rolename"] == name), "f")
        ws.append([
            name,
            role_ja.get(name, ""),
            role_memo.get(name, ""),
            n_codes,
            "○" if is_sys == "t" else "",
        ])
    for col, width in zip("ABCDE", (22, 22, 52, 12, 14)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    # ── 権限マトリクス ────────────────────────────────────────────────────
    ws = wb.create_sheet("権限マトリクス")
    ws.append(["業務領域（権限コード）", "対象アプリ"] + [
        f"{role_ja.get(r, r)}\n({r})" for r in role_names
    ])
    style_header(ws, 1, 2 + len(role_names))
    for code in codes:
        row = [f"{perm_ja.get(code, code)}\n{code}", "\n".join(code_apps.get(code, []))]
        for name in role_names:
            cell = cells.get((name, code))
            if not cell:
                row.append("")
                continue
            letters = sort_letters(cell["letters"])
            scopes = cell["scopes"] - {"ALL"}
            row.append(f"{letters} ({'/'.join(sorted(scopes))})" if scopes else letters)
        ws.append(row)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 34
    for i in range(len(role_names)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal="left" if cell.column <= 2 else "center",
            )
            if cell.column > 2 and cell.value == "◎":
                cell.fill = ADMIN_FILL
        row[0].fill = SUB_FILL
        row[0].font = Font(bold=True, size=10)
    ws.freeze_panes = "C2"

    # ── アプリと権限 ──────────────────────────────────────────────────────
    ws = wb.create_sheet("アプリと権限")
    ws.append(["区分", "操作コード", "アプリ", "パス", "必要な権限コード", "権限の表示名"])
    style_header(ws, 1, 6)
    for a in sorted(apps, key=lambda x: x["code"]):
        ws.append([
            a["category"],
            a["code"],
            a["label"],
            a["href"],
            a["perm"] or "（ログインのみ）",
            perm_ja.get(a["perm"], "") if a["perm"] else "",
        ])
    for col, width in zip("ABCDEF", (12, 12, 26, 34, 22, 22)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    # ── 検証ユーザー(dev) ─────────────────────────────────────────────────
    ws = wb.create_sheet("検証ユーザー(dev)")
    ws.append(["ユーザー名", "表示名", "割当ロール", "有効", "パスワード", "備考"])
    style_header(ws, 1, 6)
    for u in users:
        name = u["username"]
        if name.startswith("dev_"):
            pw, note = "dev2026", "app-dev.ckk-tool.co.jp（検証専用）"
        elif name.startswith("demo"):
            pw, note = "demo2026", "app-dev.ckk-tool.co.jp（検証専用）"
        else:
            pw, note = "", "SSO / 実ユーザー"
        ws.append([
            name,
            u["display_name"],
            u["roles"] or "（未割当）",
            "○" if u["is_active"] == "t" else "",
            pw,
            note,
        ])
    for col, width in zip("ABCDEF", (24, 22, 30, 8, 14, 32)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(REPO)}")
    print(f"  roles={len(role_names)} codes={len(codes)} apps={len(apps)} users={len(users)}")


if __name__ == "__main__":
    main()
